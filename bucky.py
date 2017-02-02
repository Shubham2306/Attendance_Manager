from tkinter import *
import random
from functools import partial
import openpyxl
import webbrowser


def get_my_quote(str):
    content = open("quotes.txt").read()
    content_tokens = content.split("+")
    quote = content_tokens[random.randrange(0,len(content_tokens))]
    return str+quote


def generate_center(master):
    master.withdraw()
    master.update_idletasks()
    x = (master.winfo_screenwidth() - master.winfo_reqwidth()) / 2
    y = ((master.winfo_screenheight() - master.winfo_reqheight()) / 2) - 50
    master.geometry("+%d+%d" % (x, y))
    master.deiconify()


def click_function(event):
    str = "\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n"
    nextQuote = get_my_quote(str)
    background_label.config(text=nextQuote)


def color_config(widget, color, event):
    widget.configure(background=color)
    widget.config(cursor="hand2")


def open_file():
    #print("called")
    file = "output.txt"
    webbrowser.open(file)


def click_save_btn(master, row_btn, col_btn):
    row_value = row_btn.get()
    col_value = col_btn.get()
    file = open("setting.txt","w")
    file.write(row_value+" ")
    file.write(col_value)
    file.close()
    pop_up_msg("saved")


def settings_window():
    root = Tk()
    content = open(r"setting.txt").read()
    content_tokens = content.split(" ")
    saved_row = content_tokens[0]
    saved_column = content_tokens[1]
    root.title("Attendance Manager | Settings")
    root.configure(background="ghost white")
    root.geometry('400x250')
    top_frame = Frame(root)
    top_frame.configure(background="ghost white")
    top_frame.pack(pady=20)
    title_label = Label(top_frame, text="Where does the ID start in your excel sheet?", bg="ghost white",
                        font=("Arial", 12, "bold"))
    title_label.pack()
    bottom_frame = Frame(root)
    bottom_frame.configure(background="ghost white")
    bottom_frame.pack(pady=10)
    label_row = Label(bottom_frame, text="Row", bg="ghost white", font=("Arial", 10))
    label_row.grid(row=0, column=0, sticky=W)
    row_input = Entry(bottom_frame, width=10, bd=1)
    row_input.focus()
    row_input.grid(row=0, column=1, sticky=W, padx=10)
    row_input.delete(0, END)
    row_input.insert(0, saved_row)
    label_col = Label(bottom_frame, text="Column", bg="ghost white", font=("Arial", 10))
    label_col.grid(row=0, column=2, sticky=W)
    col_input = Entry(bottom_frame, width=10, bd=1)
    col_input.grid(row=0, column=3, sticky=W, padx=10)
    col_input.delete(0, END)
    col_input.insert(0, saved_column)
    btn_frame = Frame(root)
    btn_frame.pack(anchor=E, padx=50, pady=20)
    save_btn = Button(btn_frame, text="Save", bg="dark slate gray", fg="white", font=("Arial", 10), width=10, command=lambda: click_save_btn(root, row_input, col_input))
    save_btn.bind("<Enter>", partial(color_config, save_btn, "slate gray"))
    save_btn.bind("<Leave>", partial(color_config, save_btn, "dark slate gray"))
    save_btn.pack(expand=YES, fill=X)
    generate_center(root)
    root.iconbitmap(r'icon.ico')
    root.resizable(width=False, height=False)
    root.mainloop()


def pop_up_msg(msg):
    popup = Tk()
    popup.configure(background="ghost white")
    popup.geometry('280x180')
    top_frame = Frame(popup)
    top_frame.configure(background="ghost white")
    top_frame.pack(pady=10)
    bottom_frame = Frame(popup)
    bottom_frame.configure(background="ghost white")
    bottom_frame.pack(pady=10)

    if msg == "op_true":

        popup.wm_title("Alert")
        label = Label(top_frame, text="Few IDs were not found!!", bg="ghost white", fg="red", font=("Arial", 12),
                      pady=10)
        label.grid(row=1, columnspan=2)
        btn1 = Button(bottom_frame, text="Download IDs", bg="dark slate gray", fg="white", font=("Arial", 10),
                      command=open_file)
        btn1.bind("<Enter>", partial(color_config, btn1, "slate gray"))
        btn1.bind("<Leave>", partial(color_config, btn1, "dark slate gray"))
        btn1.grid(row=0, column=0, padx=5)
        btn2 = Button(bottom_frame, text="Close", bg="dark slate gray", fg="white", font=("Arial", 10),
                      command=popup.destroy, width=10)
        btn2.bind("<Enter>", partial(color_config, btn2, "slate gray"))
        btn2.bind("<Leave>", partial(color_config, btn2, "dark slate gray"))
        btn2.grid(row=0, column=1, padx=5)

    elif msg == "op_false":
        popup.wm_title("Done")
        label = Label(top_frame, text="Sucessfully Updated!!", fg="green", bg="ghost white", font=("Arial", 12, "bold"),
                      pady=10)
        label.grid(row=1, columnspan=2)
        btn2 = Button(bottom_frame, text="Close", bg="dark slate gray", fg="white", font=("Arial", 10),
                      command=popup.destroy)
        btn2.bind("<Enter>", partial(color_config, btn2, "slate gray"))
        btn2.bind("<Leave>", partial(color_config, btn2, "dark slate gray"))
        btn2.grid(row=0, column=0, padx=5)

    elif msg == "saved":
        popup.wm_title("Saved")
        label = Label(top_frame, text="Index saved successfully!!", fg="green", bg="ghost white", font=("Arial",12, "bold"), pady=10)
        label.grid(row=1, columnspan=2)
        btn2 = Button(bottom_frame, text="Close", bg="dark slate gray", fg="white", font=("Arial",10), command=popup.destroy)
        btn2.bind("<Enter>", partial(color_config, btn2, "slate gray"))
        btn2.bind("<Leave>", partial(color_config, btn2, "dark slate gray"))
        btn2.grid(row=0, column=0, padx=5)

    generate_center(popup)
    popup.resizable(width=False, height=False)
    popup.iconbitmap(r'icon.ico')
    popup.mainloop()


def function_called(btn_widget, read_widget_for_xlsx, read_widget_for_txt, read_widget_for_date, read_widget_trend, xlsx_msg, txt_msg, date_msg, trend_msg):
    print("trying to update--")
    flag = 0
    file_path_xlsx = read_widget_for_xlsx.get()
    file_path_txt = read_widget_for_txt.get()
    date_ip = read_widget_for_date.get()
    trend = read_widget_trend.get()
    xlsx_msg.config(font=("Arial",10,"bold"))
    txt_msg.config(font=("Arial",10,"bold"))
    date_msg.config(font=("Arial",10,"bold"))
    trend_msg.config(font=("Arial",10,"bold"))
    if file_path_xlsx == "":
        flag = 1
        xlsx_msg.config(text="Invalid")
        xlsx_msg.config(foreground="red")
    else:
        xlsx_msg.config(text="")
        flag = 0;
    if file_path_txt == "":
        flag = 1
        txt_msg.config(text="Invalid")
        txt_msg.config(foreground="red")
    else:
        flag = 0;
        txt_msg.config(text="")
    if date_ip == "":
        flag = 1
        date_msg.config(text="Invalid")
        date_msg.config(foreground="red")
    else:
        flag = 0
        date_msg.config(text="")
    if trend == "" or len(trend) != 9:
        flag = 1
        trend_msg.config(text="Invalid")
        trend_msg.config(foreground="red")
    else:
        flag = 0
        trend_msg.config(text="")
    if flag == 0:
        try:
            book = openpyxl.load_workbook(file_path_xlsx)
        except FileNotFoundError:
            xlsx_msg.config(text="File not found!!")
            xlsx_msg.config(foreground="red")
            flag=1
        try:
            content = open(file_path_txt).read()
            content_tokens = content.split(" ")
        except FileNotFoundError:
            txt_msg.config(text="File not found!!")
            txt_msg.config(foreground="red")
            flag=1
    if flag == 0:
        freshers_list = []
        backlogger_list = []
        NNF = []

        trend = int(trend)
        for each in range(0, len(content_tokens)):
            if int(content_tokens[each]) < trend:
                freshers_list.append(trend + int(content_tokens[each]))
            else:
                backlogger_list.append(int(content_tokens[each]))

        freshers_list.sort()
        backlogger_list.sort()

        freshers_list_len = len(freshers_list)
        backlogger_list_len = len(backlogger_list)

        attendance_marked = 0
        count = 0
        sheet = book.active

        row_file = open("setting.txt").read()
        row_file_tokens = row_file.split(" ")

        val_row = int(row_file_tokens[0])
        val_col = int(row_file_tokens[1])
        ptr = val_col

        while sheet.cell(row=val_row-1, column=val_col).value != None :
            val_col += 1

        sheet.cell(row=val_row-1, column=val_col).value = date_ip

        if int(sheet.cell(row=val_row, column=ptr).value) < trend:
            print("less than trend")
            #starts with backloggers
            row_counter=3

            for _ in range(0, len(backlogger_list)):
                if int(sheet.cell(row=row_counter, column=ptr).value) == backlogger_list[count]:
                    sheet.cell(row=row_counter, column=val_col).value = 1
                    attendance_marked += 1
                    row_counter += 1
                    count += 1
                elif int(sheet.cell(row=row_counter, column=ptr).value) > backlogger_list[count] and int(sheet.cell(row=row_counter, column=2).value) < trend:
                    NNF.append(sheet.cell(row=row_counter, column=2).value)
                    count += 1
                else:
                    sheet.cell(row=row_counter, column=val_col).value = 0
                    row_counter += 1

            #starts with freshers

            count=0

            for _ in range(0, len(freshers_list)):
                if int(sheet.cell(row=row_counter, column=ptr).value) == freshers_list[count]:
                    sheet.cell(row=row_counter, column=val_col).value = 1
                    attendance_marked += 1
                    row_counter += 1
                    count += 1
                elif int(sheet.cell(row=row_counter, column=ptr).value) > freshers_list[count]:
                    NNF.append(sheet.cell(row=row_counter, column=2).value)
                    count += 1
                else:
                    sheet.cell(row=row_counter, column=val_col).value = 0
                    row_counter += 1

        else:
            row_counter = 3
            for _ in range(0,len(freshers_list)):
                if int(sheet.cell(row=row_counter, column=ptr).value) == freshers_list[count]:
                    sheet.cell(row=row_counter, column=val_col).value = 1
                    attendance_marked += 1
                    row_counter += 1
                    count += 1
                elif int(sheet.cell(row=row_counter, column=ptr).value) > freshers_list[count]:
                    NNF.append(sheet.cell(row=row_counter, column=val_col).value)
                    count += 1
                else:
                    sheet.cell(row=row_counter, column=val_col).value = 0
                    row_counter += 1

            count = 0

            for _ in range(0, len(backlogger_list)):
                if int(sheet.cell(row=row_counter, column=ptr).value) == backlogger_list[count]:
                    sheet.cell(row=row_counter, column=val_col).value = 1
                    attendance_marked += 1
                    row_counter += 1
                    count += 1
                elif int(sheet.cell(row=row_counter, column=ptr).value) > backlogger_list[count] and int(sheet.cell(row=row_counter, column=2).value) < trend:
                    NNF.append(sheet.cell(row=row_counter, column=2).value)
                    count += 1
                else:
                    sheet.cell(row=row_counter, column=val_col).value = 0
                    row_counter += 1

        while sheet.cell(row=row_counter, column=ptr).value != None:
             sheet.cell(row=row_counter, column=val_col).value = 0
             row_counter += 1

        sheet.cell(row=row_counter, column=val_col).value = attendance_marked

        file_name_xlsx = file_path_xlsx.split("\\")
        print("file name: "+file_name_xlsx[-1])
        book.save(file_path_xlsx)
        print(attendance_marked)
        print("not saved"+str(len(NNF)))
        if attendance_marked != freshers_list_len + backlogger_list_len:
            file = open("output.txt","w")
            file.write("Report generated on: "+date_ip+"\n*************************************************************\nFollowing students were not found in the list\n")
            for i in range(0, len(NNF)):
                file.write(str(NNF[i])+"\n")
            file.close()
            pop_up_msg("op_true")

        else:
            pop_up_msg("op_false")


def start_attendance():

    #destroy current screen
    root.destroy()

    #initialize new window
    attendance = Tk()
    attendance.title("Attendance Manager | Dashboard")
    attendance.configure(background="ghost white")


    frame_title = Frame(attendance, bg="ghost white")
    frame_form = Frame(attendance, bg="ghost white")
    frame_title.pack()
    frame_form.pack(side="bottom")

    #title
    label_title = Label(frame_title, text="Attendance Form", fg="red", bg="ghost white", font=("Arial", 16))
    label_title.pack(anchor=CENTER, pady=20)

    #form-body
    xlsx_label = Label(frame_form, text="Give absolute *.xlsx file path", bg="ghost white", font=("Arial", 10))
    xlsx_label.grid(row=0, column=0, padx=10, pady=10)

    read_xlsx = Entry(frame_form, bd=1, width=35, font=("Arial", 10))
    read_xlsx.grid(row=0, column=1, padx=10, pady=10)

    msg_xlsx = Label(frame_form, text="", width=12, bg="ghost white", font=("Arial",10))
    msg_xlsx.grid(row=0, column=2, sticky = W, padx=10, pady=10)

    read_xlsx.focus()

    txt_label = Label(frame_form, text="Give absolute *.txt file path", bg="ghost white", font=("Arial", 10))
    txt_label.grid(row=1, column=0, sticky=W, padx=10, pady=10)

    read_txt = Entry(frame_form, bd=1, width=35, font=("Arial", 10))
    read_txt.grid(row=1, column=1, sticky=W, padx=10, pady=10)

    msg_txt = Label(frame_form, text="", width=12, bg="ghost white", font=("Arial",10))
    msg_txt.grid(row=1, column=2, sticky = W, padx=10, pady=10)

    date = Label(frame_form, text = "Date", bg="ghost white", font=("Arial", 10))
    date.grid(row=3, column=0, sticky=W, padx=10, pady=10)

    read_date = Entry(frame_form, bd=1, width=35, font=("Arial",10))
    read_date.grid(row=3, column=1, sticky=W, padx=10, pady=10)

    msg_date = Label(frame_form, text="", width=12, bg="ghost white", font=("Arial",10))
    msg_date.grid(row=3, column=2, sticky = W, padx=10, pady=10)

    trend = Label(frame_form, text="Base(9-digit)", bg="ghost white", font=("Arial", 10))
    trend.grid(row=4, column=0, sticky=W, padx=10, pady=10)

    read_trend = Entry(frame_form, bd=1, width=35, font=("Arial", 10))
    read_trend.grid(row=4, column=1, sticky=W, padx=10, pady=10)

    msg_trend = Label(frame_form, text="", width=12, bg="ghost white", font=("Arial", 10))
    msg_trend.grid(row=4, column=2, sticky=W, padx=10, pady=10)

    #button
    submit_btn = Button(frame_form, text="Submit", width=30, font=("Arial, 10"), bg="dark slate gray", fg="white", command = lambda: function_called(submit_btn, read_xlsx, read_txt, read_date, read_trend, msg_xlsx, msg_txt, msg_date, msg_trend))
    submit_btn.grid(row=5, column=1, sticky=E, padx=10, pady=15)

    submit_btn.bind("<Enter>",partial(color_config, submit_btn, "slate gray"))
    submit_btn.bind("<Leave>", partial(color_config, submit_btn, "dark slate gray"))


    generate_center(attendance)
    attendance.resizable(width=False, height=False)
    attendance.iconbitmap(r'icon.ico')
    attendance.mainloop()


def log_result():
    file = r"output.txt"
    webbrowser.open(file)


root = Tk()
root.title("Attendance Manager")

#initializing background


photo = PhotoImage(file="background.png")

quote = get_my_quote("\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n")
background_label = Label(root, image=photo, text=quote, compound=CENTER, fg="white")
background_label.config(font=("Arial", 10, "bold"))
background_label.config(border=0)
background_label.pack(side='top', fill='both', expand='yes')



background_label.bind("<Button-1>", click_function)
menuBar = Menu(root)
attendance_photo = PhotoImage(file="attendance.png")
subMenu = Menu(menuBar, tearoff=0)
menuBar.add_cascade(label="Start", menu=subMenu)
subMenu.add_command(label="  Update File", image = attendance_photo, compound=LEFT, command=start_attendance)

setting_icon = PhotoImage(file="settings.png")
search_icon = PhotoImage(file="search.png")
view = Menu(menuBar, tearoff=0)
menuBar.add_cascade(label="View", menu=view)
view.add_cascade(label="  Log Results", image=search_icon, compound=LEFT, command=open_file)
view.add_cascade(label="  Settings", image=setting_icon, compound=LEFT, command=settings_window)
root.config(menu=menuBar)

#generating center

generate_center(root)
root.iconbitmap(r'icon.ico')
root.resizable(width=False, height=False)

root.mainloop()
